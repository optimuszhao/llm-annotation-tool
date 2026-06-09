from __future__ import annotations

import math
import re
from io import BytesIO
from typing import Any, Iterable, Optional
from uuid import uuid4

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from backend.database import decode_json, encode_json, get_db, insert_many, now_iso

ROLE_RESULT_KEY = "角色标注结果"
ROLE_ANSWER_KEY = "角色标注答案"
ROLE_RESULT_COLUMN_SEPARATOR = "."
MODEL_RESULT_INTERNAL_KEYS = {ROLE_RESULT_KEY, ROLE_ANSWER_KEY}
PREVIEW_LIMIT = 120
ANALYSIS_RESULT_COLUMN_PREFIX = "分析结果｜"


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _preview(value: Any, limit: int = PREVIEW_LIMIT) -> Any:
    if value is None:
        return ""
    text = encode_json(value) if isinstance(value, (dict, list)) else str(value)
    return text if len(text) <= limit else f"{text[:limit]}..."


def _is_large_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (dict, list)):
        return True
    text = str(value)
    return len(text) > PREVIEW_LIMIT or "\n" in text


def build_row_preview_payload(row_data: dict[str, Any]) -> tuple[str, str]:
    preview_data = {key: _preview(value) for key, value in row_data.items()}
    large_fields = [
        key
        for key, value in row_data.items()
        if _is_large_value(value)
    ]
    return encode_json(preview_data), encode_json(large_fields)


def _normalize_answer(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text.startswith("mock_"):
        text = text[5:]
    return "".join(text.split())


def _answer_polarity(value: Any) -> Optional[bool]:
    normalized = _normalize_answer(value)
    positive_values = {"1", "true", "yes", "y", "positive", "pos", "是", "有", "正", "正例", "阳性", "命中"}
    negative_values = {"0", "false", "no", "n", "negative", "neg", "否", "无", "负", "负例", "阴性", "未命中"}
    if normalized in positive_values:
        return True
    if normalized in negative_values:
        return False
    return None


def _infer_import_status(row_data: dict, field_mapping: Optional[dict]) -> str:
    human_column = (field_mapping or {}).get("human_answer_column") or ""
    model_column = (field_mapping or {}).get("model_answer_column") or ""
    if not human_column or not model_column:
        return "未标注"
    if human_column not in row_data or model_column not in row_data:
        return "未标注"

    human_value = row_data.get(human_column)
    model_value = row_data.get(model_column)
    human_text = _normalize_answer(human_value)
    model_text = _normalize_answer(model_value)
    if not human_text or not model_text:
        return "未标注"

    human_positive = _answer_polarity(human_value)
    model_positive = _answer_polarity(model_value)
    if human_positive is not None and model_positive is not None:
        if human_positive and model_positive:
            return "TP"
        if not human_positive and not model_positive:
            return "TN"
        if not human_positive and model_positive:
            return "FP"
        return "FN"

    return "TP" if human_text == model_text else "FP"


async def import_excel_files(scene_id: str, files: list[UploadFile]) -> list[dict]:
    if not files:
        raise HTTPException(status_code=400, detail="请选择 Excel 文件")

    imported: list[dict] = []
    with get_db() as conn:
        scene = conn.execute("SELECT * FROM scenes WHERE id=?", (scene_id,)).fetchone()
        if not scene:
            raise HTTPException(status_code=404, detail="场景不存在")
        field_mapping = conn.execute(
            "SELECT human_answer_column, model_answer_column FROM field_mappings WHERE scene_id=?",
            (scene_id,),
        ).fetchone()

        for file in files:
            content = await file.read()
            try:
                workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"{file.filename} 解析失败：{exc}") from exc

            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                raise HTTPException(status_code=400, detail=f"{file.filename} 没有表头")

            column_mappings, columns = _excel_column_mappings(header_row)
            if not columns:
                raise HTTPException(status_code=400, detail=f"{file.filename} 表头为空")

            dataset_id = f"dataset_{uuid4().hex[:12]}"
            timestamp = now_iso()
            data_rows: list[tuple[str, str, int, str, str, str, str, str]] = []
            next_row_index = 1
            for row in rows:
                row_data = {
                    column: _cell_value(row[index] if index < len(row) else "")
                    for index, column in column_mappings
                }
                if all(value == "" for value in row_data.values()):
                    continue
                preview_data, large_fields = build_row_preview_payload(row_data)
                data_rows.append(
                    (
                        f"row_{uuid4().hex[:16]}",
                        dataset_id,
                        next_row_index,
                        encode_json(row_data),
                        preview_data,
                        large_fields,
                        _infer_import_status(row_data, field_mapping),
                        timestamp,
                    )
                )
                next_row_index += 1

            conn.execute(
                """
                INSERT INTO datasets(id, scene_id, name, file_name, row_count, column_schema, created_at)
                VALUES(?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    dataset_id,
                    scene_id,
                    file.filename or dataset_id,
                    file.filename,
                    len(data_rows),
                    encode_json(columns),
                    timestamp,
                ),
            )
            insert_many(
                conn,
                f"""
                INSERT INTO {scene['data_table_name']}(id, dataset_id, row_index, raw_data, preview_data, large_fields, annotation_status, created_at)
                VALUES(?, ?, ?, ?, ?, ?, ?, ?)
                """,
                data_rows,
            )
            imported.append(
                {
                    "id": dataset_id,
                    "scene_id": scene_id,
                    "name": file.filename or dataset_id,
                    "file_name": file.filename,
                    "row_count": len(data_rows),
                    "column_schema": columns,
                    "created_at": timestamp,
                }
            )
    return imported


def list_datasets(scene_id: Optional[str] = None) -> list[dict]:
    with get_db() as conn:
        if scene_id:
            rows = conn.execute(
                "SELECT * FROM datasets WHERE scene_id=? ORDER BY created_at DESC",
                (scene_id,),
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM datasets ORDER BY created_at DESC").fetchall()
    for row in rows:
        row["column_schema"] = decode_json(row["column_schema"], [])
    return rows


def delete_dataset(dataset_id: str) -> dict:
    with get_db() as conn:
        dataset = conn.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
        if not dataset:
            raise HTTPException(status_code=404, detail="数据集不存在")
        scene = conn.execute("SELECT * FROM scenes WHERE id=?", (dataset["scene_id"],)).fetchone()
        if not scene:
            raise HTTPException(status_code=404, detail="场景不存在")
        table_name = scene["data_table_name"]
        conn.execute(f"DELETE FROM {table_name} WHERE dataset_id=?", (dataset_id,))
        conn.execute("DELETE FROM row_analysis_history WHERE dataset_id=?", (dataset_id,))
        conn.execute("DELETE FROM datasets WHERE id=?", (dataset_id,))
    return {"ok": True, "id": dataset_id}


def get_dataset_row(dataset_id: str, row_id: str, scheme_id: str = "") -> dict:
    with get_db() as conn:
        dataset, _scene, row = _load_dataset_row(conn, dataset_id, row_id, scheme_id)
        model_result_columns = _model_result_columns(conn, dataset_id)
        columns = _sync_model_result_columns(conn, dataset, model_result_columns)
    return _format_row(row, columns, preview=False, scheme_view=bool(scheme_id))


def get_dataset_row_field(dataset_id: str, row_id: str, column: str, scheme_id: str = "") -> dict:
    with get_db() as conn:
        dataset, _scene, row = _load_dataset_row(conn, dataset_id, row_id, scheme_id)
        columns = _sync_model_result_columns(conn, dataset)
        analysis_columns = _analysis_result_columns(conn, dataset_id)
        analysis_method = _analysis_method_by_column(analysis_columns).get(column)
        if analysis_method:
            latest_analysis = conn.execute(
                """
                SELECT analysis_data
                FROM row_analysis_history
                WHERE dataset_id=? AND row_id=? AND method_name=?
                ORDER BY created_at DESC, rowid DESC
                LIMIT 1
                """,
                (dataset_id, row_id, analysis_method),
            ).fetchone()
            value = decode_json(latest_analysis["analysis_data"], {}) if latest_analysis else {}
            text = encode_json(value)
            return {
                "row_id": row_id,
                "column": column,
                "value": value,
                "size": len(text),
                "exists": bool(latest_analysis),
            }
    raw_data = decode_json(row["raw_data"], {})
    if scheme_id:
        model_result = decode_json(row.get("scheme_model_result"), {})
        raw_data = {**raw_data, **flatten_model_result_for_display(model_result)}
    elif column not in raw_data:
        model_result = decode_json(row.get("model_result"), {})
        raw_data = {**raw_data, **flatten_model_result_for_display(model_result)}
    value = raw_data.get(column, "")
    text = value if isinstance(value, str) else encode_json(value)
    return {
        "row_id": row_id,
        "column": column,
        "value": value,
        "size": len(str(text)),
        "exists": column in raw_data or column in columns,
    }


def update_dataset_row(dataset_id: str, row_id: str, payload: dict, scheme_id: str = "") -> dict:
    raw_data = payload.get("raw_data") if isinstance(payload, dict) else None
    if raw_data is None:
        raw_data = payload
    if not isinstance(raw_data, dict):
        raise HTTPException(status_code=400, detail="行数据必须是 JSON 对象")

    raw_data = {
        str(key): value
        for key, value in raw_data.items()
        if key not in {"row_id", "row_index", "状态", "model_result", "analysis_data", "rendered_prompt"}
    }
    timestamp = now_iso()
    with get_db() as conn:
        dataset, scene, row = _load_dataset_row(conn, dataset_id, row_id)
        table_name = scene["data_table_name"]
        field_mapping = conn.execute(
            "SELECT human_answer_column, model_answer_column FROM field_mappings WHERE scene_id=?",
            (scene["id"],),
        ).fetchone()
        model_result = decode_json(row.get("model_result"), {})
        analysis_data = decode_json(row.get("analysis_data"), {})
        for key, value in model_result.items():
            raw_data.setdefault(key, value)
        if analysis_data and "分析数据" not in raw_data:
            raw_data["分析数据"] = analysis_data
        columns = decode_json(dataset["column_schema"], [])
        for key in raw_data:
            if key not in columns:
                columns.append(key)
        preview_data, large_fields = build_row_preview_payload(raw_data)
        annotation_status = _infer_import_status(raw_data, field_mapping)
        conn.execute(
            "UPDATE datasets SET column_schema=? WHERE id=?",
            (encode_json(columns), dataset_id),
        )
        conn.execute(
            f"""
            UPDATE {table_name}
            SET raw_data=?, preview_data=?, large_fields=?, annotation_status=?, updated_at=?
            WHERE id=? AND dataset_id=?
            """,
            (encode_json(raw_data), preview_data, large_fields, annotation_status, timestamp, row_id, dataset_id),
        )
        if scheme_id:
            _update_latest_scheme_row_status(conn, dataset_id, row_id, scheme_id, raw_data, field_mapping, timestamp)
        dataset, _scene, row = _load_dataset_row(conn, dataset_id, row_id, scheme_id)
    return _format_row(row, columns, preview=False, scheme_view=bool(scheme_id))


def _update_latest_scheme_row_status(
    conn,
    dataset_id: str,
    row_id: str,
    scheme_id: str,
    raw_data: dict,
    field_mapping: Optional[dict],
    timestamp: str,
) -> None:
    latest = conn.execute(
        """
        SELECT task_row.id, task_row.model_result
        FROM annotation_task_rows task_row
        JOIN annotation_tasks task ON task.id=task_row.task_id
        WHERE task.dataset_id=? AND task.scheme_id=? AND task_row.row_id=?
        ORDER BY COALESCE(task_row.finished_at, task_row.updated_at, task_row.created_at) DESC
        LIMIT 1
        """,
        (dataset_id, scheme_id, row_id),
    ).fetchone()
    if not latest:
        return
    scheme_result = decode_json(latest.get("model_result"), {})
    status = _infer_import_status({**raw_data, **scheme_result}, field_mapping)
    conn.execute(
        "UPDATE annotation_task_rows SET status=?, updated_at=? WHERE id=?",
        (status, timestamp, latest["id"]),
    )


def _excel_column_mappings(header_row: tuple[Any, ...]) -> tuple[list[tuple[int, str]], list[str]]:
    mappings: list[tuple[int, str]] = []
    columns: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(header_row):
        column = str(value).strip() if value not in (None, "") else ""
        if not column:
            continue
        count = seen.get(column, 0) + 1
        seen[column] = count
        unique_column = column if count == 1 else f"{column}_{count}"
        while unique_column in columns:
            count += 1
            unique_column = f"{column}_{count}"
        mappings.append((index, unique_column))
        columns.append(unique_column)
    return mappings, columns


def delete_dataset_row(dataset_id: str, row_id: str) -> dict:
    with get_db() as conn:
        _dataset, scene, _row = _load_dataset_row(conn, dataset_id, row_id)
        table_name = scene["data_table_name"]
        conn.execute(f"DELETE FROM {table_name} WHERE id=? AND dataset_id=?", (row_id, dataset_id))
        conn.execute("DELETE FROM annotation_task_rows WHERE row_id=?", (row_id,))
        conn.execute("DELETE FROM row_analysis_history WHERE dataset_id=? AND row_id=?", (dataset_id, row_id))
        total = conn.execute(
            f"SELECT COUNT(*) AS total FROM {table_name} WHERE dataset_id=?",
            (dataset_id,),
        ).fetchone()["total"]
        conn.execute("UPDATE datasets SET row_count=? WHERE id=?", (total, dataset_id))
    return {"ok": True, "id": row_id, "dataset_id": dataset_id, "row_count": total}


def delete_dataset_rows(dataset_id: str, row_ids: list[str]) -> dict:
    ids = [row_id for row_id in row_ids if row_id]
    if not ids:
        raise HTTPException(status_code=400, detail="请选择要删除的行")
    with get_db() as conn:
        dataset = conn.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
        if not dataset:
            raise HTTPException(status_code=404, detail="数据集不存在")
        scene = conn.execute("SELECT * FROM scenes WHERE id=?", (dataset["scene_id"],)).fetchone()
        if not scene:
            raise HTTPException(status_code=404, detail="场景不存在")
        table_name = scene["data_table_name"]
        placeholders = ", ".join(["?"] * len(ids))
        existing = conn.execute(
            f"SELECT id FROM {table_name} WHERE dataset_id=? AND id IN ({placeholders})",
            [dataset_id, *ids],
        ).fetchall()
        existing_ids = [row["id"] for row in existing]
        if existing_ids:
            existing_placeholders = ", ".join(["?"] * len(existing_ids))
            conn.execute(
                f"DELETE FROM {table_name} WHERE dataset_id=? AND id IN ({existing_placeholders})",
                [dataset_id, *existing_ids],
            )
            conn.execute(
                f"DELETE FROM annotation_task_rows WHERE row_id IN ({existing_placeholders})",
                existing_ids,
            )
            conn.execute(
                f"DELETE FROM row_analysis_history WHERE dataset_id=? AND row_id IN ({existing_placeholders})",
                [dataset_id, *existing_ids],
            )
        total = conn.execute(
            f"SELECT COUNT(*) AS total FROM {table_name} WHERE dataset_id=?",
            (dataset_id,),
        ).fetchone()["total"]
        conn.execute("UPDATE datasets SET row_count=? WHERE id=?", (total, dataset_id))
    return {
        "ok": True,
        "dataset_id": dataset_id,
        "deleted_count": len(existing_ids),
        "row_ids": existing_ids,
        "row_count": total,
    }


def get_dataset_rows(
    dataset_id: str,
    page: int = 1,
    page_size: int = 50,
    search: str = "",
    search_column: str = "",
    search_empty: bool = False,
    scheme_id: str = "",
    statuses: Optional[list[str]] = None,
    favorite_only: bool = False,
    sort_field: str = "",
    sort_dir: str = "asc",
    root_cause_value: str = "",
    root_cause_positive: Optional[list[str]] = None,
    root_cause_negative: Optional[list[str]] = None,
) -> dict:
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)
    offset = (page - 1) * page_size

    with get_db() as conn:
        dataset = conn.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
        if not dataset:
            raise HTTPException(status_code=404, detail="数据集不存在")
        scene = conn.execute("SELECT * FROM scenes WHERE id=?", (dataset["scene_id"],)).fetchone()
        table_name = scene["data_table_name"]
        field_mapping = conn.execute(
            "SELECT root_cause_column FROM field_mappings WHERE scene_id=?",
            (dataset["scene_id"],),
        ).fetchone()
        root_cause_column = (field_mapping["root_cause_column"] if field_mapping else "").strip()
        model_result_columns = _model_result_columns(conn, dataset_id)
        columns = _sync_model_result_columns(conn, dataset, model_result_columns)
        analysis_columns = _analysis_result_columns(conn, dataset_id)
        analysis_method_by_column = _analysis_method_by_column(analysis_columns)
        all_columns = [*columns, *[item["column"] for item in analysis_columns]]
        scheme_view = bool(scheme_id)
        base_alias = "d" if scheme_view else ""
        status_expr = "COALESCE(latest.scheme_status, '未标注')" if scheme_view else "annotation_status"
        sort_field = sort_field.strip()
        sort_dir = "DESC" if str(sort_dir).lower() == "desc" else "ASC"
        where = f"{base_alias + '.' if base_alias else ''}dataset_id=?"
        params: list[Any] = [dataset_id]
        search_text = search.strip()
        search_column = search_column.strip()
        if (search_text or search_empty) and search_column in {"display_index", "row_index", "序号"}:
            index_column = f"{base_alias + '.' if base_alias else ''}row_index"
            if search_text:
                row_indexes = _parse_row_index_filter(search_text)
                if row_indexes:
                    placeholders = ", ".join(["?"] * len(row_indexes))
                    where += f" AND {index_column} IN ({placeholders})"
                    params.extend(row_indexes)
                else:
                    where += " AND 1=0"
            elif search_empty:
                where += " AND 1=0"
        elif search_text and search_column == "状态":
            where += f" AND {status_expr} LIKE ?"
            params.append(f"%{search_text}%")
        elif search_empty and search_column == "状态":
            where += f" AND ({status_expr} IS NULL OR TRIM(COALESCE(CAST({status_expr} AS TEXT), ''))='')"
        elif (search_text or search_empty) and search_column and search_column in analysis_method_by_column:
            analysis_expr = _latest_analysis_expr(base_alias)
            if search_text:
                where += f" AND COALESCE(CAST({analysis_expr} AS TEXT), '') LIKE ?"
                params.extend([dataset_id, analysis_method_by_column[search_column], f"%{search_text}%"])
            if search_empty:
                where += f" AND TRIM(COALESCE(CAST({analysis_expr} AS TEXT), ''))=''"
                params.extend([dataset_id, analysis_method_by_column[search_column]])
        elif (search_text or search_empty) and search_column and search_column in columns:
            clauses: list[str] = []
            if search_text:
                if scheme_view:
                    clauses.append(
                        """
                        (
                          COALESCE(CAST(json_extract(d.raw_data, ?) AS TEXT), '') LIKE ?
                          OR COALESCE(CAST(json_extract(latest.scheme_model_result, ?) AS TEXT), '') LIKE ?
                        )
                        """
                    )
                    params.extend([
                        _json_column_path(search_column),
                        f"%{search_text}%",
                        _model_result_json_path(search_column),
                        f"%{search_text}%",
                    ])
                else:
                    clauses.append("COALESCE(CAST(json_extract(raw_data, ?) AS TEXT), '') LIKE ?")
                    params.extend([_json_column_path(search_column), f"%{search_text}%"])
            if search_empty:
                if scheme_view:
                    clauses.append(
                        """
                        (
                          (
                            json_type(latest.scheme_model_result, ?) IS NOT NULL
                            AND TRIM(COALESCE(CAST(json_extract(latest.scheme_model_result, ?) AS TEXT), ''))=''
                          )
                          OR
                          (
                            json_type(latest.scheme_model_result, ?) IS NULL
                            AND (
                              json_type(d.raw_data, ?) IS NULL
                              OR TRIM(COALESCE(CAST(json_extract(d.raw_data, ?) AS TEXT), ''))=''
                            )
                          )
                        )
                        """
                    )
                    params.extend([
                        _model_result_json_path(search_column),
                        _model_result_json_path(search_column),
                        _model_result_json_path(search_column),
                        _json_column_path(search_column),
                        _json_column_path(search_column),
                    ])
                else:
                    clauses.append(
                        """
                        (
                          json_type(raw_data, ?) IS NULL
                          OR TRIM(COALESCE(CAST(json_extract(raw_data, ?) AS TEXT), ''))=''
                        )
                        """
                    )
                    params.extend([_json_column_path(search_column), _json_column_path(search_column)])
            where += f" AND ({' OR '.join(clauses)})"
        elif search_text:
            if scheme_view:
                where += " AND (d.raw_data LIKE ? OR COALESCE(latest.scheme_model_result, '') LIKE ?)"
                params.extend([f"%{search_text}%", f"%{search_text}%"])
            else:
                where += " AND raw_data LIKE ?"
                params.append(f"%{search_text}%")
        status_values = _normalize_status_filters(statuses)
        if status_values:
            concrete_statuses = [status for status in status_values if status != "未标注"]
            clauses: list[str] = []
            if "未标注" in status_values:
                clauses.append(f"({status_expr} IS NULL OR {status_expr}='' OR {status_expr}='未标注')")
            if concrete_statuses:
                placeholders = ", ".join(["?"] * len(concrete_statuses))
                clauses.append(f"{status_expr} IN ({placeholders})")
                params.extend(concrete_statuses)
            where += f" AND ({' OR '.join(clauses)})"
        root_cause_text = root_cause_value.strip()
        if root_cause_text and root_cause_column:
            raw_path = _json_column_path(root_cause_column)
            result_path = _model_result_json_path(root_cause_column)
            if scheme_view:
                where += """
                    AND TRIM(COALESCE(
                      CAST(json_extract(latest.scheme_model_result, ?) AS TEXT),
                      CAST(json_extract(d.model_result, ?) AS TEXT),
                      CAST(json_extract(d.raw_data, ?) AS TEXT),
                      ''
                    ))=?
                """
                params.extend([result_path, result_path, raw_path, root_cause_text])
            else:
                where += """
                    AND TRIM(COALESCE(
                      CAST(json_extract(model_result, ?) AS TEXT),
                      CAST(json_extract(raw_data, ?) AS TEXT),
                      ''
                    ))=?
                """
                params.extend([result_path, raw_path, root_cause_text])
        root_cause_filter_clauses: list[str] = []
        root_cause_filter_params: list[Any] = []
        for polarity, values in (
            ("positive", root_cause_positive or []),
            ("negative", root_cause_negative or []),
        ):
            names = [str(value).strip() for value in values if str(value).strip()]
            if not names:
                continue
            placeholders = ", ".join(["?"] * len(names))
            root_cause_filter_clauses.append(
                f"""
                {base_alias + '.' if base_alias else ''}id IN (
                  SELECT row_id
                  FROM root_cause_row_links
                  WHERE dataset_id=?
                    AND scheme_id=?
                    AND polarity=?
                    AND name IN ({placeholders})
                )
                """
            )
            root_cause_filter_params.extend([dataset_id, scheme_id or "", polarity, *names])
        if root_cause_filter_clauses:
            where += f" AND ({' OR '.join(root_cause_filter_clauses)})"
            params.extend(root_cause_filter_params)
        if favorite_only:
            where += f" AND {base_alias + '.' if base_alias else ''}is_favorite=1"
        order_params: list[Any] = []
        if sort_field == "状态":
            order_by = f"{status_expr} {sort_dir}, {base_alias + '.' if base_alias else ''}row_index ASC"
        elif sort_field in analysis_method_by_column:
            order_by = f"COALESCE(CAST({_latest_analysis_expr(base_alias)} AS TEXT), '') {sort_dir}, {base_alias + '.' if base_alias else ''}row_index ASC"
            order_params.extend([dataset_id, analysis_method_by_column[sort_field]])
        elif sort_field in columns:
            prefix = "d." if scheme_view else ""
            if scheme_view:
                order_by = (
                    "COALESCE("
                    "CAST(json_extract(latest.scheme_model_result, ?) AS TEXT), "
                    "CAST(json_extract(d.raw_data, ?) AS TEXT), "
                    f"'' ) {sort_dir}, d.row_index ASC"
                )
                order_params.extend([_model_result_json_path(sort_field), _json_column_path(sort_field)])
            else:
                order_by = f"COALESCE(CAST(json_extract({prefix}raw_data, ?) AS TEXT), '') {sort_dir}, {prefix}row_index ASC"
                order_params.append(_json_column_path(sort_field))
        else:
            order_by = f"{base_alias + '.' if base_alias else ''}row_index ASC"

        if scheme_view:
            latest_cte = _latest_scheme_rows_cte()
            latest_params = [dataset_id, scheme_id]
            total = conn.execute(
                f"""
                {latest_cte}
                SELECT COUNT(*) AS total
                FROM {table_name} d
                LEFT JOIN latest_scheme_rows latest ON latest.row_id=d.id
                WHERE {where}
                """,
                [*latest_params, *params],
            ).fetchone()["total"]
            rows = conn.execute(
                f"""
                {latest_cte}
                SELECT
                    d.id,
                    d.row_index,
                    COALESCE(NULLIF(d.preview_data, '{{}}'), d.raw_data) AS preview_data,
                    d.large_fields,
                    d.is_favorite,
                    d.annotation_status,
                    d.model_result,
                    d.analysis_data,
                    d.rendered_prompt,
                    latest.scheme_status,
                    latest.scheme_model_result,
                    latest.scheme_analysis_data,
                    latest.scheme_rendered_prompt
                FROM {table_name} d
                LEFT JOIN latest_scheme_rows latest ON latest.row_id=d.id
                WHERE {where}
                ORDER BY {order_by}
                LIMIT ? OFFSET ?
                """,
                [*latest_params, *params, *order_params, page_size, offset],
            ).fetchall()
        else:
            total = conn.execute(
                f"SELECT COUNT(*) AS total FROM {table_name} WHERE {where}",
                params,
            ).fetchone()["total"]
            rows = conn.execute(
                f"""
                SELECT
                    id,
                    row_index,
                    COALESCE(NULLIF(preview_data, '{{}}'), raw_data) AS preview_data,
                    large_fields,
                    is_favorite,
                    annotation_status,
                    model_result
                FROM {table_name}
                WHERE {where}
                ORDER BY {order_by}
                LIMIT ? OFFSET ?
                """,
                [*params, *order_params, page_size, offset],
            ).fetchall()

        data = [_format_row(row, columns, preview=True, scheme_view=bool(scheme_id)) for row in rows]
        for display_index, item in enumerate(data, start=offset + 1):
            item["display_index"] = item.get("row_index") or display_index
        _append_latest_analysis_results(conn, dataset_id, data, analysis_columns)

    return {
        "data": data,
        "total": total,
        "last_row": total,
        "page": page,
        "page_size": page_size,
        "last_page": max(math.ceil(total / page_size), 1),
        "columns": all_columns,
        "model_result_columns": model_result_columns,
    }


def _json_column_path(column: str) -> str:
    escaped = column.replace("\\", "\\\\").replace('"', '\\"')
    return f'$."{escaped}"'


def _parse_row_index_filter(value: str) -> list[int]:
    indexes: list[int] = []
    seen: set[int] = set()
    for part in re.split(r"[,，\\s]+", value.strip()):
        if not part:
            continue
        try:
            index = int(part)
        except ValueError:
            continue
        if index < 1 or index in seen:
            continue
        seen.add(index)
        indexes.append(index)
    return indexes


def _model_result_json_path(column: str) -> str:
    role_name, result_key = _split_role_result_column(column)
    if role_name and result_key:
        parts = [ROLE_RESULT_KEY, role_name, *[part for part in result_key.split(ROLE_RESULT_COLUMN_SEPARATOR) if part]]
        return _json_path_from_parts(parts)
    return _json_column_path(column)


def _json_path_from_parts(parts: list[str]) -> str:
    return "$" + "".join(f'."{_escape_json_path_key(str(part))}"' for part in parts if part)


def _escape_json_path_key(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def _split_role_result_column(column: str) -> tuple[str, str]:
    text = str(column or "")
    if ROLE_RESULT_COLUMN_SEPARATOR not in text:
        return "", ""
    role_name, result_key = text.split(ROLE_RESULT_COLUMN_SEPARATOR, 1)
    return role_name.strip(), result_key.strip()


def _model_result_columns(conn, dataset_id: str) -> list[str]:
    rows = []
    dataset = conn.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
    if dataset:
        scene = conn.execute("SELECT * FROM scenes WHERE id=?", (dataset["scene_id"],)).fetchone()
        if scene:
            rows.extend(conn.execute(
                f"""
                SELECT model_result
                FROM {scene['data_table_name']}
                WHERE dataset_id=?
                  AND json_valid(model_result)
                  AND TRIM(COALESCE(model_result, '')) NOT IN ('', '{{}}')
                ORDER BY row_index ASC
                """,
                (dataset_id,),
            ).fetchall())
    rows.extend(conn.execute(
        """
        SELECT task_row.model_result
        FROM annotation_task_rows task_row
        JOIN annotation_tasks task ON task.id=task_row.task_id
        WHERE task.dataset_id=?
          AND json_valid(task_row.model_result)
          AND TRIM(COALESCE(task_row.model_result, '')) NOT IN ('', '{}')
        ORDER BY COALESCE(task_row.finished_at, task_row.updated_at, task_row.created_at) DESC
        """,
        (dataset_id,),
    ).fetchall())
    columns: list[str] = []
    for row in rows:
        columns.extend(model_result_display_columns(decode_json(row["model_result"], {})))
    return _dedupe_columns(columns)


def model_result_display_columns(model_result: dict) -> list[str]:
    return list(flatten_model_result_for_display(model_result).keys())


def _role_result_column(role_name: str, field_name: str) -> str:
    return f"{role_name}{ROLE_RESULT_COLUMN_SEPARATOR}{field_name}"


def _dedupe_columns(columns: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for column in columns:
        if not column or column in seen:
            continue
        seen.add(column)
        result.append(column)
    return result


def _sync_model_result_columns(conn, dataset: dict, model_result_columns: Optional[list[str]] = None) -> list[str]:
    current = decode_json(dataset["column_schema"], [])
    columns = [
        column
        for column in current
        if column and column not in MODEL_RESULT_INTERNAL_KEYS
    ]
    for column in model_result_columns if model_result_columns is not None else _model_result_columns(conn, dataset["id"]):
        if column and column not in columns:
            columns.append(column)
    if columns != current:
        conn.execute(
            "UPDATE datasets SET column_schema=? WHERE id=?",
            (encode_json(columns), dataset["id"]),
        )
    return columns


def reindex_dataset_rows(dataset_id: str) -> dict:
    with get_db() as conn:
        dataset = conn.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
        if not dataset:
            raise HTTPException(status_code=404, detail="数据集不存在")
        scene = conn.execute("SELECT * FROM scenes WHERE id=?", (dataset["scene_id"],)).fetchone()
        if not scene:
            raise HTTPException(status_code=404, detail="场景不存在")
        table_name = scene["data_table_name"]
        rows = conn.execute(
            f"""
            SELECT id, row_index
            FROM {table_name}
            WHERE dataset_id=?
            ORDER BY row_index ASC, id ASC
            """,
            (dataset_id,),
        ).fetchall()
        timestamp = now_iso()
        updated_count = 0
        for index, row in enumerate(rows, start=1):
            if int(row["row_index"] or 0) == index:
                continue
            conn.execute(
                f"UPDATE {table_name} SET row_index=?, updated_at=? WHERE id=? AND dataset_id=?",
                (index, timestamp, row["id"], dataset_id),
            )
            conn.execute(
                "UPDATE annotation_task_rows SET row_index=?, updated_at=? WHERE row_id=?",
                (index, timestamp, row["id"]),
            )
            updated_count += 1
        conn.execute("UPDATE datasets SET row_count=? WHERE id=?", (len(rows), dataset_id))
    return {
        "ok": True,
        "dataset_id": dataset_id,
        "row_count": len(rows),
        "updated_count": updated_count,
        "start_index": 1 if rows else 0,
        "end_index": len(rows),
    }


def export_dataset_rows(dataset_id: str) -> dict:
    with get_db() as conn:
        dataset = conn.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
        if not dataset:
            raise HTTPException(status_code=404, detail="数据集不存在")
        scene = conn.execute("SELECT * FROM scenes WHERE id=?", (dataset["scene_id"],)).fetchone()
        if not scene:
            raise HTTPException(status_code=404, detail="场景不存在")
        table_name = scene["data_table_name"]
        columns = decode_json(dataset["column_schema"], [])
        rows = conn.execute(
            f"""
            SELECT id, row_index, raw_data, annotation_status, model_result, analysis_data, rendered_prompt
            FROM {table_name}
            WHERE dataset_id=?
            ORDER BY row_index ASC
            """,
            (dataset_id,),
        ).fetchall()
    return {
        "dataset": {
            "id": dataset["id"],
            "name": dataset["name"],
            "row_count": dataset["row_count"],
            "columns": columns,
        },
        "rows": [_format_row(row, columns, preview=False) for row in rows],
    }


def _load_dataset_row(conn, dataset_id: str, row_id: str, scheme_id: str = "") -> tuple[dict, dict, dict]:
    dataset = conn.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
    if not dataset:
        raise HTTPException(status_code=404, detail="数据集不存在")
    scene = conn.execute("SELECT * FROM scenes WHERE id=?", (dataset["scene_id"],)).fetchone()
    if not scene:
        raise HTTPException(status_code=404, detail="场景不存在")
    if scheme_id:
        row = conn.execute(
            f"""
            {_latest_scheme_rows_cte()}
            SELECT
                d.id,
                d.row_index,
                d.raw_data,
                d.preview_data,
                d.large_fields,
                d.is_favorite,
                d.annotation_status,
                d.model_result,
                d.analysis_data,
                d.rendered_prompt,
                latest.scheme_status,
                latest.scheme_model_result,
                latest.scheme_analysis_data,
                latest.scheme_rendered_prompt
            FROM {scene['data_table_name']} d
            LEFT JOIN latest_scheme_rows latest ON latest.row_id=d.id
            WHERE d.id=? AND d.dataset_id=?
            """,
            (dataset_id, scheme_id, row_id, dataset_id),
        ).fetchone()
    else:
        row = conn.execute(
            f"""
            SELECT id, row_index, raw_data, preview_data, large_fields, is_favorite, annotation_status, model_result, analysis_data, rendered_prompt
            FROM {scene['data_table_name']}
            WHERE id=? AND dataset_id=?
            """,
            (row_id, dataset_id),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="数据行不存在")
    return dataset, scene, row


def _normalize_status_filters(statuses: Optional[list[str]]) -> list[str]:
    values: list[str] = []
    for status in statuses or []:
        values.extend(part.strip() for part in str(status).split(",") if part.strip())
    return list(dict.fromkeys(values))


def _analysis_result_columns(conn, dataset_id: str) -> list[dict[str, str]]:
    rows = conn.execute(
        """
        SELECT
            method_name,
            COALESCE(NULLIF(method_label, ''), method_name) AS method_label,
            MAX(created_at) AS latest_at
        FROM row_analysis_history
        WHERE dataset_id=?
          AND TRIM(COALESCE(method_name, ''))!=''
        GROUP BY method_name
        ORDER BY latest_at DESC, method_label ASC
        """,
        (dataset_id,),
    ).fetchall()
    label_counts: dict[str, int] = {}
    for row in rows:
        label = row["method_label"] or row["method_name"]
        label_counts[label] = label_counts.get(label, 0) + 1
    result = []
    for row in rows:
        label = row["method_label"] or row["method_name"]
        title = label if label_counts.get(label, 0) == 1 else f"{label}（{row['method_name']}）"
        result.append({
            "column": f"{ANALYSIS_RESULT_COLUMN_PREFIX}{title}",
            "method_name": row["method_name"],
            "method_label": label,
        })
    return result


def _analysis_method_by_column(analysis_columns: list[dict[str, str]]) -> dict[str, str]:
    return {item["column"]: item["method_name"] for item in analysis_columns}


def _latest_analysis_expr(base_alias: str = "") -> str:
    row_expr = f"{base_alias}.id" if base_alias else "id"
    return f"""
        (
          SELECT h.analysis_data
          FROM row_analysis_history h
          WHERE h.dataset_id=? AND h.row_id={row_expr} AND h.method_name=?
          ORDER BY h.created_at DESC, h.rowid DESC
          LIMIT 1
        )
    """


def _append_latest_analysis_results(conn, dataset_id: str, data: list[dict], analysis_columns: list[dict[str, str]]) -> None:
    if not data or not analysis_columns:
        return
    row_ids = [row["row_id"] for row in data if row.get("row_id")]
    if not row_ids:
        return
    placeholders = ", ".join(["?"] * len(row_ids))
    rows = conn.execute(
        f"""
        SELECT row_id, method_name, analysis_data
        FROM (
          SELECT
            row_id,
            method_name,
            analysis_data,
            ROW_NUMBER() OVER (
              PARTITION BY row_id, method_name
              ORDER BY created_at DESC, rowid DESC
            ) AS rn
          FROM row_analysis_history
          WHERE dataset_id=? AND row_id IN ({placeholders})
        )
        WHERE rn=1
        """,
        [dataset_id, *row_ids],
    ).fetchall()
    column_by_method = {item["method_name"]: item["column"] for item in analysis_columns}
    values_by_row: dict[str, dict[str, str]] = {}
    for row in rows:
        column = column_by_method.get(row["method_name"])
        if not column:
            continue
        values_by_row.setdefault(row["row_id"], {})[column] = _preview(decode_json(row.get("analysis_data"), {}))
    for item in data:
        item.update(values_by_row.get(item.get("row_id"), {}))


def _format_row(row: dict, columns: list[str], preview: bool, scheme_view: bool = False) -> dict:
    raw_data = decode_json(row.get("preview_data") if preview else row.get("raw_data"), {})
    if scheme_view:
        model_result = decode_json(row.get("scheme_model_result"), {})
        display_result = flatten_model_result_for_display(model_result)
        if preview:
            display_result = {key: _preview(value) for key, value in display_result.items()}
        raw_data = {**raw_data, **display_result}
        status = row.get("scheme_status") or "未标注"
    else:
        model_result = decode_json(row.get("model_result"), {})
        display_result = flatten_model_result_for_display(model_result)
        if preview:
            display_result = {key: _preview(value) for key, value in display_result.items()}
        raw_data = {**raw_data, **display_result}
        status = row.get("annotation_status") or raw_data.get("状态") or raw_data.get("status") or "未标注"
    large_fields = set(decode_json(row.get("large_fields"), []) or [])
    item = {
        "row_id": row["id"],
        "row_index": row["row_index"],
        "状态": status,
        "is_favorite": bool(row.get("is_favorite")),
        "收藏": "是" if row.get("is_favorite") else "否",
    }
    infer_large_fields = preview and not large_fields
    for column in columns:
        value = raw_data.get(column, "")
        if infer_large_fields and _is_large_value(value):
            large_fields.add(column)
        item[column] = _preview(value) if preview else value
    item["is_favorite"] = bool(row.get("is_favorite"))
    item["收藏"] = "是" if row.get("is_favorite") else "否"
    if preview:
        item["__large_fields"] = sorted(large_fields)
    if not preview:
        analysis_data = decode_json(row.get("scheme_analysis_data") if scheme_view else row.get("analysis_data"), {})
        item["model_result"] = model_result
        item["analysis_data"] = analysis_data
        item["rendered_prompt"] = (row.get("scheme_rendered_prompt") if scheme_view else row.get("rendered_prompt")) or ""
        if analysis_data and "分析数据" not in item:
            item["分析数据"] = analysis_data
    return item


def _flatten_role_model_result(model_result: dict) -> dict[str, Any]:
    role_results = model_result.get(ROLE_RESULT_KEY) if isinstance(model_result, dict) else {}
    if not isinstance(role_results, dict):
        return {}
    flattened: dict[str, Any] = {}
    for role_name, result in role_results.items():
        if not isinstance(result, dict):
            continue
        for key, value in _flatten_leaf_values(result).items():
            flattened[_role_result_column(str(role_name), str(key))] = value
    return flattened


def flatten_model_result_for_display(model_result: dict) -> dict[str, Any]:
    if not isinstance(model_result, dict):
        return {}
    top_level_result = _flatten_leaf_values({
        key: value
        for key, value in model_result.items()
        if key not in MODEL_RESULT_INTERNAL_KEYS
    })
    role_result = _flatten_role_model_result(model_result)
    return {**top_level_result, **role_result}


def _flatten_leaf_values(value: Any, prefix: str = "") -> dict[str, Any]:
    if isinstance(value, dict):
        flattened: dict[str, Any] = {}
        for key, item in value.items():
            if not key:
                continue
            next_prefix = f"{prefix}{ROLE_RESULT_COLUMN_SEPARATOR}{key}" if prefix else str(key)
            flattened.update(_flatten_leaf_values(item, next_prefix))
        return flattened
    return {prefix: value} if prefix else {}


def update_dataset_row_favorite(dataset_id: str, row_id: str, is_favorite: bool) -> dict:
    with get_db() as conn:
        dataset, scene, row = _load_dataset_row(conn, dataset_id, row_id)
        timestamp = now_iso()
        conn.execute(
            f"UPDATE {scene['data_table_name']} SET is_favorite=?, updated_at=? WHERE id=? AND dataset_id=?",
            (1 if is_favorite else 0, timestamp, row_id, dataset_id),
        )
    return {
        "row_id": row_id,
        "is_favorite": bool(is_favorite),
        "收藏": "是" if is_favorite else "否",
    }


def update_dataset_rows_favorite(dataset_id: str, payload: dict) -> dict:
    row_ids = [str(row_id) for row_id in payload.get("row_ids") or [] if row_id]
    is_favorite = bool(payload.get("is_favorite"))
    favorite_only = bool(payload.get("favorite_only", False))
    with get_db() as conn:
        dataset = conn.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
        if not dataset:
            raise HTTPException(status_code=404, detail="数据集不存在")
        scene = conn.execute("SELECT * FROM scenes WHERE id=?", (dataset["scene_id"],)).fetchone()
        if not scene:
            raise HTTPException(status_code=404, detail="场景不存在")
        table_name = scene["data_table_name"]
        where = "dataset_id=?"
        params: list[Any] = [dataset_id]
        if row_ids:
            placeholders = ", ".join(["?"] * len(row_ids))
            where += f" AND id IN ({placeholders})"
            params.extend(row_ids)
        if favorite_only:
            where += " AND is_favorite=1"
        rows = conn.execute(f"SELECT id FROM {table_name} WHERE {where}", params).fetchall()
        target_ids = [row["id"] for row in rows]
        if target_ids:
            placeholders = ", ".join(["?"] * len(target_ids))
            conn.execute(
                f"UPDATE {table_name} SET is_favorite=?, updated_at=? WHERE id IN ({placeholders})",
                [1 if is_favorite else 0, now_iso(), *target_ids],
            )
    return {
        "updated_count": len(target_ids),
        "row_ids": target_ids,
        "is_favorite": is_favorite,
        "收藏": "是" if is_favorite else "否",
    }


def clear_dataset_rows_favorite(dataset_id: str, payload: dict) -> dict:
    row_ids = [str(row_id) for row_id in payload.get("row_ids") or [] if row_id]
    favorite_only = bool(payload.get("favorite_only", True))
    with get_db() as conn:
        dataset = conn.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
        if not dataset:
            raise HTTPException(status_code=404, detail="数据集不存在")
        scene = conn.execute("SELECT * FROM scenes WHERE id=?", (dataset["scene_id"],)).fetchone()
        if not scene:
            raise HTTPException(status_code=404, detail="场景不存在")
        table_name = scene["data_table_name"]
        where = "dataset_id=?"
        params: list[Any] = [dataset_id]
        if row_ids:
            placeholders = ", ".join(["?"] * len(row_ids))
            where += f" AND id IN ({placeholders})"
            params.extend(row_ids)
        if favorite_only:
            where += " AND is_favorite=1"
        rows = conn.execute(f"SELECT id FROM {table_name} WHERE {where}", params).fetchall()
        target_ids = [row["id"] for row in rows]
        if target_ids:
            placeholders = ", ".join(["?"] * len(target_ids))
            conn.execute(
                f"UPDATE {table_name} SET is_favorite=0, updated_at=? WHERE id IN ({placeholders})",
                [now_iso(), *target_ids],
            )
    return {
        "updated_count": len(target_ids),
        "row_ids": target_ids,
    }


def _latest_scheme_rows_cte() -> str:
    return """
    WITH latest_scheme_rows AS (
      SELECT row_id, scheme_status, scheme_model_result, scheme_analysis_data, scheme_rendered_prompt
      FROM (
        SELECT
          task_row.row_id,
          task_row.status AS scheme_status,
          task_row.model_result AS scheme_model_result,
          task_row.analysis_data AS scheme_analysis_data,
          task_row.rendered_prompt AS scheme_rendered_prompt,
          ROW_NUMBER() OVER (
            PARTITION BY task_row.row_id
            ORDER BY COALESCE(task_row.finished_at, task_row.updated_at, task_row.created_at) DESC
          ) AS rn
        FROM annotation_task_rows task_row
        JOIN annotation_tasks task ON task.id=task_row.task_id
        WHERE task.dataset_id=? AND task.scheme_id=?
      )
      WHERE rn=1
    )
    """
